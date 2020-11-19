import openpyxl as xl
from openpyxl.chart import (
    LineChart,
    Reference
)

def importInfectionRates(infections, dates, hcd):

    select = 0

    while select not in ['1','2']:
        select = input('1. Save as...\n2. Back\n')
        if select == '1':
            filename = input('Save as: ')
        if select == '2':
            return 

    wb = xl.Workbook()
    wb.save(f'./output/{filename}.xlsx')
    wb = xl.load_workbook(f'./output/{filename}.xlsx')
    sheet = wb.active

    header = ['Infections', 'Date']

    # Write the headers
    for col in range(1,len(header)+1):
        header_row = sheet.cell(1, col)
        header_row.value = header[col-1]
        
    bottom_row = row = 2
    for i in infections:
        infections_row = sheet.cell(row, 1)
        infections_row.value = i
        row += 1
        bottom_row += 1

    row = 2
    for i in dates:
        date_row = sheet.cell(row, 2)
        date_row.value = i
        row += 1
        
    chart = LineChart()
    chart.title = f'Confirmed cases, {hcd}, {dates[0]} - {dates[-1]}'
    chart.style = 13
    chart.y_axis.title = 'Infections'
    chart.x_axis.title = 'Date'

    values = Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=bottom_row)
    dates = Reference(sheet, min_col=2, min_row=2, max_col=2, max_row=bottom_row)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(dates)

    sheet.add_chart(chart, "D2")

    wb.save(f'./output/{filename}.xlsx')


def importHospitalCases(totalHospitalised, inWard, inIcu, dead, dates, hospital):

    select = 0

    while select not in ['1','2']:
        select = input('1. Save as...\n2. Back\n')
        if select == '1':
            filename = input('Save as: ')
        if select == '2':
            return 

    wb = xl.Workbook()
    wb.save(f'./output/{filename}.xlsx')
    wb = xl.load_workbook(f'./output/{filename}.xlsx')
    sheet = wb.active

    header = ['Total hospitalised', 'In ward', 'In icu', 'Dead', 'Date']

    # Write the headers
    for col in range(1,len(header)+1):
        header_row = sheet.cell(1, col)
        header_row.value = header[col-1]
        
    bottom_row = row = 2
    for i in totalHospitalised:
        totalHospitalised_row = sheet.cell(row, 1)
        totalHospitalised_row.value = i
        row += 1
        bottom_row += 1

    row = 2
    for i in inWard:
        inWard_row = sheet.cell(row, 2)
        inWard_row.value = i
        row += 1

    row = 2
    for i in inIcu:
        inIcu_row = sheet.cell(row, 3)
        inIcu_row.value = i
        row += 1

    row = 2
    for i in dead:
        dead_row = sheet.cell(row, 4)
        dead_row.value = i
        row += 1

    row = 2
    for i in dates:
        date_row = sheet.cell(row, 5)
        date_row.value = i
        row += 1
        
    chart = LineChart()
    chart.title = f'Hospital data, {hospital}, {dates[0]} - {dates[-1]}'
    chart.style = 13
    chart.y_axis.title = 'Infections'
    chart.x_axis.title = 'Date'

    values = Reference(sheet, min_col=1, min_row=1, max_col=4, max_row=bottom_row)
    dates = Reference(sheet, min_col=5, min_row=2, max_col=5, max_row=bottom_row)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(dates)
    
    sheet.add_chart(chart, "G2")

    wb.save(f'./output/{filename}.xlsx')

